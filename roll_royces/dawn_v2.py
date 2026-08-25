import re
import os
import pandas as pd
from openpyxl import load_workbook

# Dataset configurations
DATASETS = {
    "BaTriTemp": {
        "missing_percentage": [8, 16, 24, 40, 56],
        "seeds": [6479, 2843, 5388, 9485, 1857],
    },
    "BaTriHumidity": {
        "missing_percentage": [8, 16, 24, 40, 56],
        "seeds": [2843, 5543, 2818, 2025, 9999],
    },
}

# Other configurations
COMBINATION_MODE = ["meow", "data_per"]
MODELS = ["LinearRegression", "KNeighborsRegressor", "RandomForestRegressor", "DecisionTreeRegressor", "SVR", "AdaBoostRegressor", "CNN1D", "attention", "multi_attention"]
QUERY_CONFIG = {
    "directory_format": "{_}_{dataset}_{missing_percentage}_{seed}_{combination_mode}",
    "col": {
        "target": "Model",
        "value": MODELS
    },
    "row": {
        "only": ["Similarity", "NMAE", "RMSE", "R2", "FSD", "FB", "FA2"]
    },
    "out_delimiter": "\t",
}

# Ensure the output directory exists
os.makedirs("./metrics", exist_ok=True)

# Extract query details
query = QUERY_CONFIG["directory_format"]
query_keys = re.findall(r"\{([^}]+)\}", query)

# Build regex pattern
pattern = rf"^{query}$"
for key in query_keys:
    pattern = pattern.replace(f"{{{key}}}", f"(.+?)")

# List all directories in results
directories = os.listdir("results")

# Get the row clause
row = QUERY_CONFIG["row"]
row_only = row.get("only", [])

# Function to append or create a sheet
def append_or_create_sheet(dataframe, file_path, sheet_name):
    if os.path.exists(file_path):
        # Load the workbook
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            workbook = writer.book
            if sheet_name in workbook.sheetnames:
                # Read existing data
                existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                # Append new data to existing data
                dataframe = pd.concat([existing_df, dataframe], ignore_index=True)
            # Overwrite the sheet with updated data
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # Write to a new file if it doesn't exist
        dataframe.to_excel(file_path, sheet_name=sheet_name, index=False)

# Merge cell if the value is the same
def merge_cell(output_file, sheet_name, column):
    wb = load_workbook(output_file)
    ws = wb[sheet_name]
    current_value = None
    start_row = None  # Start row for merging

    for row in range(2, ws.max_row + 1):
        new_value = ws[f"{column}{row}"].value
        if new_value != current_value:
            # Merge the previous range if applicable
            if current_value is not None and start_row is not None and row - start_row > 1:
                ws.merge_cells(f"{column}{start_row}:{column}{row - 1}")
            # Update start row for new value
            start_row = row
            current_value = new_value

    # Handle the final range
    if current_value is not None and start_row is not None and ws.max_row - start_row > 0:
        ws.merge_cells(f"{column}{start_row}:{column}{ws.max_row}")

    wb.save(output_file)

if __name__ == "__main__":
    for dataset, config in DATASETS.items():
        missing_percentages = config["missing_percentage"]
        seeds = config["seeds"]
        output_file = f"./metrics/metrics_summary_{dataset}.xlsx"

        # Create a dictionary to store all results for the summary
        summary_results = []

        for combination_mode in COMBINATION_MODE:
            for seed in seeds:
                all_results = []

                # Process all gaps for the current seed
                for missing_percentage in missing_percentages:
                    # Update the where clause
                    where = {
                        "dataset": dataset,
                        "seed": seed,
                        "missing_percentage": missing_percentage,
                        "combination_mode": combination_mode,
                    }

                    # Find matching directories
                    query_directories = []
                    for directory in directories:
                        _match = re.match(pattern, directory)
                        if _match:
                            match_tuple = _match.groups()
                            # Check if the where clause is satisfied
                            if all(str(match_tuple[query_keys.index(key)]) == str(value) for key, value in where.items()):
                                query_directories.append(directory)

                    # Process each matching directory
                    for directory in query_directories:
                        result_file = os.path.join("results", directory, "metrics.csv")
                        if os.path.exists(result_file):
                            df = pd.read_csv(result_file)

                            # Filter rows and columns
                            df = df[df[QUERY_CONFIG["col"]["target"]].isin(MODELS)]
                            df = df[[QUERY_CONFIG["col"]["target"]] + row_only]
                            df.insert(0, "CombinationMode", combination_mode)
                            df.insert(1, "Gap", missing_percentage)
                            all_results.append(df)
                            summary_results.append(df)  # Add to summary results

                # Combine results for all gaps in the current seed
                if all_results:
                    combined_df = pd.concat(all_results)
                    append_or_create_sheet(combined_df, output_file, sheet_name=f"Seed_{seed}")

        # Create summary sheet from all results
        if summary_results:
            final_combined_df = pd.concat(summary_results)
            
            # Calculate averages across seeds
            averages = (
                # Group by without sorting
                final_combined_df.groupby(["CombinationMode", "Gap", "Model"], sort=False)
                .mean()
                .reset_index()
            )
                        
            # Remove any existing summary sheet
            if os.path.exists(output_file):
                with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:
                    if "Summary" in writer.book.sheetnames:
                        idx = writer.book.sheetnames.index("Summary")
                        writer.book.remove(writer.book.worksheets[idx])
            
            # Write the new summary sheet
            append_or_create_sheet(averages, output_file, sheet_name="Summary")

        # Merge cells in the output file
        for seed in seeds:
            merge_cell(output_file, sheet_name=f"Seed_{seed}", column="A")
            merge_cell(output_file, sheet_name=f"Seed_{seed}", column="B")
        
        # Merge cells in summary sheet
        merge_cell(output_file, sheet_name="Summary", column="A")
        merge_cell(output_file, sheet_name="Summary", column="B")
        
        print(f"Metrics summary saved for dataset {dataset} in {output_file}")